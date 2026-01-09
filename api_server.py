"""
Badge Management API Server with Brother QL Label Printer Integration
Flask REST API for badge management and label printing - USB pyusb Connection
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import sqlite3
from datetime import datetime
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Brother QL Label Printer imports
from PIL import Image, ImageDraw, ImageFont
from brother_ql.conversion import convert
from brother_ql.backends.helpers import send
from brother_ql.raster import BrotherQLRaster

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

app = Flask(__name__)

# Configure CORS
cors_origins = os.getenv('CORS_ORIGINS', '*')
if cors_origins != '*':
    cors_origins = cors_origins.split(',')
CORS(app, origins=cors_origins)

# ==================== Configuration from Environment Variables ====================

# Database Configuration
DB_NAME = os.getenv('DB_NAME', 'badges.db')

# External API Configuration
EXTERNAL_API_URL = os.getenv('EXTERNAL_API_URL', 'http://badges.eevent.ma/api/getbadges')
EXTERNAL_PRINCIPAUX_API_URL = os.getenv('EXTERNAL_PRINCIPAUX_API_URL', 'https://eevent.ma/api/inscritslemm')

# Brother QL Printer Configuration
PRINTER_MODEL = os.getenv('PRINTER_MODEL', 'QL-810W')
PRINTER_BACKEND = os.getenv('PRINTER_BACKEND', 'pyusb')

# USB Configuration
PRINTER_USB_VENDOR_ID = os.getenv('PRINTER_USB_VENDOR_ID', '0x04f9')
PRINTER_USB_PRODUCT_ID = os.getenv('PRINTER_USB_PRODUCT_ID', '0x209c')
PRINTER_IDENTIFIER = f"usb://{PRINTER_USB_VENDOR_ID}:{PRINTER_USB_PRODUCT_ID}"

# Label Configuration
LABEL_SIZE = os.getenv('LABEL_SIZE', '29x90')
LABEL_ROTATE = os.getenv('LABEL_ROTATE', '90')
LABEL_CUT = os.getenv('LABEL_CUT', 'True').lower() == 'true'

# Server Configuration
FLASK_HOST = os.getenv('FLASK_HOST', '0.0.0.0')
FLASK_PORT = int(os.getenv('FLASK_PORT', '5000'))
FLASK_DEBUG = os.getenv('FLASK_DEBUG', 'True').lower() == 'true'

# Detect operating system
IS_WINDOWS = sys.platform.startswith('win')

# ==================== Database Helper Functions ====================

def get_db_connection():
    """Create database connection"""
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize database"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            prenom TEXT NOT NULL,
            valide INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS print_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            printed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')
    
    conn.commit()
    conn.close()

# ==================== Font Helper Functions ====================

def get_font_path(font_name, font_size):
    """Get font path based on operating system"""
    font = None
    
    # Check if custom font path is provided in environment
    custom_font = os.getenv('WINDOWS_FONT_PATH' if IS_WINDOWS else 'LINUX_FONT_PATH')
    if custom_font and os.path.exists(custom_font):
        try:
            font = ImageFont.truetype(custom_font, font_size)
            print(f"Using custom font: {custom_font}")
            return font
        except Exception as e:
            print(f"Could not load custom font {custom_font}: {e}")
    
    if IS_WINDOWS:
        # Windows font paths
        windows_fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
        font_paths = [
            os.path.join(windows_fonts_dir, 'arial.ttf'),
            os.path.join(windows_fonts_dir, 'arialbd.ttf'),
            os.path.join(windows_fonts_dir, 'calibri.ttf'),
            os.path.join(windows_fonts_dir, 'calibrib.ttf'),
            os.path.join(windows_fonts_dir, 'segoeui.ttf'),
            os.path.join(windows_fonts_dir, 'segoeuib.ttf'),
        ]
    else:
        # Linux/Unix font paths
        font_paths = [
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
            '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
            'arial.ttf',
            'calibri.ttf',
            'DejaVuSans.ttf',
        ]
    
    # Try each font path
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font = ImageFont.truetype(font_path, font_size)
                print(f"Using font: {font_path}")
                return font
            except Exception as e:
                print(f"Could not load font {font_path}: {e}")
                continue
    
    # Fallback to default font
    print("Warning: No TrueType font found, using default font")
    return ImageFont.load_default()

# ==================== USB Printer Detection ====================

def detect_brother_printer():
    """Detect connected Brother printer via USB"""
    try:
        import usb.core
        
        # Brother's Vendor ID from environment or default
        BROTHER_VENDOR_ID = int(PRINTER_USB_VENDOR_ID, 16)
        
        # Find all Brother devices
        devices = usb.core.find(find_all=True, idVendor=BROTHER_VENDOR_ID)
        
        printers = []
        for device in devices:
            usb_id = f"usb://0x{device.idVendor:04x}:0x{device.idProduct:04x}"
            printers.append(usb_id)
            print(f"Found Brother printer: {usb_id}")
        
        return printers
    except ImportError:
        print("Warning: pyusb not installed. Run: pip install pyusb")
        return []
    except Exception as e:
        print(f"Error detecting printer: {e}")
        return []

# ==================== Brother QL Label Printer Functions ====================

def create_label_image(first_name, last_name):
    """Create label image for Brother QL printer"""
    # Label dimensions for 29mm x 90mm (in pixels, as expected by brother_ql)
    label_width = 991   # Effective printable width for 90mm
    label_height = 306  # 29mm at 300 DPI

    # Create a grayscale image
    image = Image.new("L", (label_width, label_height), "white")
    draw = ImageDraw.Draw(image)

    # Combine first and last name
    full_name = f"{first_name} {last_name}"

    # Maximum dimensions for text (95% of width, 90% of height for larger text)
    max_text_width = int(label_width * 0.95)    # ≈941px
    max_text_height = int(label_height * 0.9)   # ≈275px

    # Start with a large font size and scale down
    font_size = 120  # Start large for bigger text
    font = None
    
    while font_size > 20:  # Minimum font size
        font = get_font_path("arial.ttf", font_size)
        
        if font:
            text_bbox = draw.textbbox((0, 0), full_name, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]

            # Check if text fits within max dimensions
            if text_width <= max_text_width and text_height <= max_text_height:
                break
            font_size -= 5  # Reduce font size and try again
        else:
            break

    # Use minimum font size if text still doesn't fit
    if font_size <= 20:
        font_size = 20
        font = get_font_path("arial.ttf", font_size)

    # Calculate text size and position for centering
    text_bbox = draw.textbbox((0, 0), full_name, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]

    # Center text horizontally
    x = (label_width - text_width) // 2

    # Center text vertically with adjustment for font metrics
    if hasattr(font, 'getmetrics'):
        try:
            ascent, descent = font.getmetrics()
            text_visual_height = ascent - descent
        except:
            text_visual_height = text_height
    else:
        text_visual_height = text_height
    
    y = (label_height - text_visual_height) // 2 - text_bbox[1]

    # Draw text
    draw.text((x, y), full_name, fill="black", font=font)

    print(f"Using font size: {font_size}pt for '{full_name}'")
    return image

def print_to_brother_ql(first_name, last_name, 
                       printer_identifier=None, 
                       model=None):
    """Print label to Brother QL printer via USB"""
    try:
        # Use provided parameters or fall back to environment config
        if printer_identifier is None:
            printer_identifier = PRINTER_IDENTIFIER
        if model is None:
            model = PRINTER_MODEL
            
        # Create label image
        image = create_label_image(first_name, last_name)
        if image is None:
            error_msg = "Failed to create label image"
            print(error_msg)
            return {"status": "error", "message": error_msg}

        # Convert and send to printer
        print(f"Printing to {printer_identifier} using {PRINTER_BACKEND} backend...")
        
        qlr = BrotherQLRaster(model)
        qlr.exception_on_warning = True
        
        instructions = convert(
            qlr=qlr,
            images=[image],
            label=LABEL_SIZE,
            rotate=LABEL_ROTATE,
            threshold=70.0,
            dither=False,
            compress=False,
            red=False,      # Set to True for black/red labels
            dpi_600=False,
            hq=True,
            cut=LABEL_CUT
        )

        # Send to printer via USB using pyusb
        send(
            instructions=instructions,
            printer_identifier=printer_identifier,
            backend_identifier=PRINTER_BACKEND,
            blocking=True
        )
        
        print(f"✓ Label printed successfully for '{first_name} {last_name}'!")
        return {"status": "success", "message": "Label printed successfully"}
        
    except Exception as e:
        error_msg = f"Error printing label: {str(e)}"
        print(error_msg)
        print(f"Printer Identifier: {printer_identifier}")
        print(f"Backend: {PRINTER_BACKEND}")
        print(f"Model: {model}")
        
        # Additional troubleshooting info
        if "No backend available" in str(e):
            error_msg += "\n\nTroubleshooting: pyusb backend not available. Install with: pip install pyusb"
        elif "Access denied" in str(e) or "Permission denied" in str(e):
            if IS_WINDOWS:
                error_msg += "\n\nTroubleshooting: USB access denied. Install WinUSB driver using Zadig."
            else:
                error_msg += "\n\nTroubleshooting: USB permission denied. Add user to lp group or create udev rule."
        elif "No such device" in str(e):
            error_msg += "\n\nTroubleshooting: Printer not found. Check USB connection and run printer detection."
        
        return {"status": "error", "message": error_msg}

# ==================== Excel Functions ====================

def create_excel_export(badges_data):
    """Create Excel file from badges data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Badges"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['ID', 'Prénom', 'Nom', 'Validé', 'Date de création', 'Dernière modification', 'Source']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data
    for badge in badges_data:
        row = [
            badge.get('id', ''),
            badge.get('prenom', ''),
            badge.get('nom', ''),
            'Oui' if badge.get('valide') == 1 else 'Non',
            badge.get('created_at', ''),
            badge.get('updated_at', ''),
            badge.get('source', 'local')
        ]
        ws.append(row)
        
        # Apply border to all cells in the row
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Adjust column widths
    column_widths = [8, 20, 20, 10, 20, 20, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==================== API Routes ====================

@app.route('/api/import-excel-template', methods=['GET'])
def download_template():
    """Download Excel import template"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Import Badges"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Prénom', 'Nom', 'Validé']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add example rows
        examples = [
            ['Mohamed', 'Alami', 'Oui'],
            ['Fatima', 'Bennani', 'Non'],
            ['Ahmed', 'El Idrissi', 'Oui'],
        ]
        
        for example in examples:
            ws.append(example)
            # Style example rows
            for cell in ws[ws.max_row]:
                cell.fill = example_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add instructions
        ws['A6'] = 'INSTRUCTIONS:'
        ws['A6'].font = Font(bold=True, size=11)
        
        ws['A7'] = '1. Remplissez les colonnes Prénom et Nom (obligatoires)'
        ws['A8'] = '2. Colonne Validé: "Oui" ou "Non" (optionnel, par défaut "Non")'
        ws['A9'] = '3. Supprimez les lignes d\'exemple avant l\'import'
        ws['A10'] = '4. Sauvegardez le fichier et importez-le dans l\'application'
        
        for row in range(7, 11):
            ws[f'A{row}'].font = Font(italic=True, size=10)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='badges_import_template.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
@app.route('/')
def index():
    """API information endpoint"""
    return jsonify({
        'name': 'Badge Management API with Brother QL Printer (USB pyusb)',
        'version': '3.0',
        'platform': 'Windows' if IS_WINDOWS else 'Linux/Unix',
        'printer': {
            'model': PRINTER_MODEL,
            'identifier': PRINTER_IDENTIFIER,
            'backend': PRINTER_BACKEND,
            'label_size': LABEL_SIZE
        },
        'endpoints': {
            'GET /api/getbadges': 'Get all badges (local + external)',
            'GET /api/getbadges/<id>': 'Get badge by ID',
            'POST /api/badges': 'Create new badge',
            'PUT /api/badges/<id>': 'Update badge',
            'DELETE /api/badges/<id>': 'Delete badge',
            'POST /print-label': 'Print badge to Brother QL printer',
            'GET /api/stats': 'Get statistics',
            'GET /api/printer/detect': 'Detect connected printers',
            'POST /api/validate/<id>': 'Validate badge',
            'GET /api/search': 'Search badges',
            'GET /api/export-excel': 'Export badges to Excel file',
            'GET /api/config': 'Get current configuration'
        }
    })

@app.route('/api/config', methods=['GET'])
def get_config():
    """Get current configuration"""
    return jsonify({
        'database': DB_NAME,
        'external_api': EXTERNAL_API_URL,
        'external_principaux': EXTERNAL_PRINCIPAUX_API_URL,
        'printer': {
            'model': PRINTER_MODEL,
            'backend': PRINTER_BACKEND,
            'identifier': PRINTER_IDENTIFIER,
            'vendor_id': PRINTER_USB_VENDOR_ID,
            'product_id': PRINTER_USB_PRODUCT_ID
        },
        'label': {
            'size': LABEL_SIZE,
            'rotate': LABEL_ROTATE,
            'cut': LABEL_CUT
        },
        'server': {
            'host': FLASK_HOST,
            'port': FLASK_PORT,
            'debug': FLASK_DEBUG
        }
    })

@app.route('/api/printer/detect', methods=['GET'])
def detect_printer():
    """Detect connected Brother printers"""
    try:
        printers = detect_brother_printer()
        
        if printers:
            return jsonify({
                'status': 'success',
                'printers': printers,
                'message': f'Found {len(printers)} Brother printer(s)',
                'current_config': PRINTER_IDENTIFIER
            })
        else:
            return jsonify({
                'status': 'warning',
                'printers': [],
                'message': 'No Brother printers detected. Check USB connection.',
                'current_config': PRINTER_IDENTIFIER
            }), 404
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

@app.route('/api/getbadges', methods=['GET'])
def get_all_badges():
    """Get all badges (combines local and external data)"""
    try:
        # Get local badges
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get filter parameters
        valide = request.args.get('valide', type=int)
        search = request.args.get('search', '')
        source = request.args.get('source', 'all')
        
        local_badges = []
        if source in ['all', 'local']:
            cursor.execute('SELECT * FROM users ORDER BY id ASC')
            users = cursor.fetchall()
            
            for user in users:
                if valide is not None and user['valide'] != valide:
                    continue
                
                if search:
                    search_lower = search.lower()
                    if (search_lower not in user['nom'].lower() and 
                        search_lower not in user['prenom'].lower() and 
                        search_lower not in str(user['id'])):
                        continue
                
                local_badges.append({
                    'id': user['id'],
                    'nom': user['nom'],
                    'prenom': user['prenom'],
                    'valide': user['valide'],
                    'created_at': user['created_at'],
                    'updated_at': user['updated_at'],
                    'source': 'local'
                })
        
        conn.close()
        
        # Get external badges qr code
        external_badges = []
        if source in ['all', 'external']:
            try:
                response = requests.get(EXTERNAL_API_URL, timeout=5)
                if response.status_code == 200:
                    external_data = response.json()
                    
                    for badge in external_data:
                        if valide is not None and badge.get('valide') != valide:
                            continue
                        
                        if search:
                            search_lower = search.lower()
                            nom = str(badge.get('nom', '')).lower()
                            prenom = str(badge.get('prenom', '')).lower()
                            email = str(badge.get('email', '')).lower()
                            badge_id = str(badge.get('id', ''))
                            
                            if search_lower not in nom and search_lower not in prenom and search_lower not in email and search_lower not in badge_id:
                                continue
                        
                        badge['source'] = 'external'
                        external_badges.append(badge)
            except requests.RequestException as e:
                print(f"Warning: Could not fetch external badges: {str(e)}")

        # Get external badges principaux
        external_badges_principaux = []
        if source in ['all', 'external_principaux']:
            try:
                response = requests.get(EXTERNAL_PRINCIPAUX_API_URL, timeout=5)
                if response.status_code == 200:
                    external_data = response.json()
                    
                    for badge in external_data['data']:
                        
                        if search:
                            search_lower = search.lower()
                            nom = str(badge.get('nom', '')).lower()
                            prenom = str(badge.get('prenom', '')).lower()
                            badge_id = str(badge.get('id', ''))
                            
                            if search_lower not in nom and search_lower not in prenom and search_lower not in badge_id:
                                continue
                        
                        badge['source'] = 'external_principaux'
                        external_badges_principaux.append(badge)
            except requests.RequestException as e:
                print(f"Warning: Could not fetch external badges: {str(e)}")
        
        # Combine results
        all_badges = local_badges + external_badges + external_badges_principaux
        # all_badges.sort(key=lambda x: x.get('id', 0))  # Ascending order (oldest first)
        
        return jsonify(all_badges)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/getbadges/<int:badge_id>', methods=['GET'])
def get_badge_by_id(badge_id):
    """Get specific badge by ID"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (badge_id,))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            return jsonify({
                'id': user['id'],
                'nom': user['nom'],
                'prenom': user['prenom'],
                'valide': user['valide'],
                'created_at': user['created_at'],
                'updated_at': user['updated_at'],
                'source': 'local'
            })
        
        try:
            response = requests.get(f"{EXTERNAL_API_URL}/{badge_id}", timeout=5)
            if response.status_code == 200:
                badge = response.json()
                badge['source'] = 'external'
                return jsonify(badge)
        except requests.RequestException:
            pass
        
        return jsonify({'error': 'Badge not found'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/print-label', methods=['POST'])
def print_label():
    """Print badge label to Brother QL printer"""
    try:
        data = request.get_json()
        
        # Support both formats
        nom = data.get('nom', data.get('last_name', ''))
        prenom = data.get('prenom', data.get('first_name', ''))
        user_id = data.get('id')
        
        if not nom or not prenom:
            return jsonify({'error': 'last_name and first_name are required'}), 400
        
        # If no ID provided, create new user
        if not user_id:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO users (nom, prenom, valide, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (nom, prenom, 1, datetime.now(), datetime.now()))
            user_id = cursor.lastrowid
            conn.commit()
            conn.close()
        
        # Print to Brother QL printer
        result = print_to_brother_ql(prenom, nom)
        
        if result['status'] == 'success':
            # Log print
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO print_logs (user_id, printed_at)
                VALUES (?, ?)
            ''', (user_id, datetime.now()))
            conn.commit()
            conn.close()
            
            return jsonify({
                'status': 'success',
                'message': f'Label printed successfully for {prenom} {nom}',
                'id': user_id
            }), 200
        else:
            return jsonify(result), 500
    
    except Exception as e:
        return jsonify({'error': str(e), 'status': 'error'}), 500

@app.route('/api/stats', methods=['GET'])
def get_statistics():
    """Get statistics"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as total FROM users')
        total = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as validated FROM users WHERE valide = 1')
        validated = cursor.fetchone()['validated']
        
        cursor.execute('SELECT COUNT(*) as printed FROM print_logs')
        printed = cursor.fetchone()['printed']
        
        conn.close()
        
        return jsonify({
            'total_badges': total,
            'validated_badges': validated,
            'total_prints': printed
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    """Import badges from Excel file"""
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
        
        # Read Excel file
        from openpyxl import load_workbook
        
        # Load workbook from uploaded file
        wb = load_workbook(filename=file, read_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).lower() if cell.value else '')
        
        # Find column indices
        try:
            prenom_idx = headers.index('prénom') if 'prénom' in headers else headers.index('prenom')
            nom_idx = headers.index('nom')
        except ValueError:
            wb.close()
            return jsonify({'error': 'Excel file must contain "Prénom" and "Nom" columns'}), 400
        
        # Optional columns
        valide_idx = headers.index('validé') if 'validé' in headers else (headers.index('valide') if 'valide' in headers else None)
        
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor()
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        # Process rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extract data
                prenom = str(row[prenom_idx]).strip() if row[prenom_idx] else ''
                nom = str(row[nom_idx]).strip() if row[nom_idx] else ''
                
                # Skip empty rows
                if not prenom or not nom:
                    skipped_count += 1
                    continue
                
                # Get valide status
                if valide_idx is not None and row[valide_idx]:
                    valide_value = str(row[valide_idx]).lower().strip()
                    valide = 1 if valide_value in ['oui', 'yes', '1', 'true', 'validé'] else 0
                else:
                    valide = 0
                
                # Check if badge already exists
                cursor.execute('''
                    SELECT id FROM users 
                    WHERE LOWER(nom) = LOWER(?) AND LOWER(prenom) = LOWER(?)
                ''', (nom, prenom))
                
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing badge
                    cursor.execute('''
                        UPDATE users 
                        SET valide = ?, updated_at = ?
                        WHERE id = ?
                    ''', (valide, datetime.now(), existing['id']))
                    skipped_count += 1
                else:
                    # Insert new badge
                    cursor.execute('''
                        INSERT INTO users (nom, prenom, valide, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (nom, prenom, valide, datetime.now(), datetime.now()))
                    imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue
        
        # Commit changes
        conn.commit()
        conn.close()
        wb.close()
        
        # Prepare response
        response = {
            'status': 'success',
            'imported': imported_count,
            'skipped': skipped_count,
            'total_processed': imported_count + skipped_count,
            'message': f'Import completed: {imported_count} new badges imported, {skipped_count} skipped (already exist or empty)'
        }
        
        if errors:
            response['errors'] = errors
            response['error_count'] = len(errors)
        
        return jsonify(response), 200
    
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """Export all badges to Excel"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users ORDER BY id ASC')
        users = cursor.fetchall()
        conn.close()
        
        badges_data = []
        for user in users:
            badges_data.append({
                'id': user['id'],
                'nom': user['nom'],
                'prenom': user['prenom'],
                'valide': user['valide'],
                'created_at': user['created_at'],
                'updated_at': user['updated_at'],
                'source': 'local'
            })
        
        excel_buffer = create_excel_export(badges_data)
        
        from flask import send_file
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'badges_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== Run Server ====================

if __name__ == '__main__':
    init_db()
    print("=" * 60)
    print("Badge Management API Server with Brother QL Printer")
    print("USB Connection via pyusb - Direct Print Only")
    print("=" * 60)
    print(f"Server starting on http://{FLASK_HOST}:{FLASK_PORT}")
    print(f"Platform: {'Windows' if IS_WINDOWS else 'Linux/Unix'}")
    print(f"Database: {DB_NAME}")
    print(f"External API: {EXTERNAL_API_URL}")
    print(f"Printer Model: {PRINTER_MODEL}")
    print(f"Printer Identifier: {PRINTER_IDENTIFIER}")
    print(f"Printer Backend: {PRINTER_BACKEND}")
    print(f"Label Size: {LABEL_SIZE}")
    print("\nPrinter endpoints:")
    print("  POST /print-label            - Print to Brother QL printer")
    print("  GET  /api/printer/detect     - Detect connected printers")
    print("  GET  /api/config             - Get current configuration")
    print("\n" + "=" * 60)
    print("USB SETUP INSTRUCTIONS:")
    print("=" * 60)
    
    if IS_WINDOWS:
        print("WINDOWS:")
        print("1. Install pyusb: pip install pyusb")
        print("2. Download Zadig: https://zadig.akeo.ie/")
        print("3. Run Zadig as Administrator")
        print("4. Options > List All Devices")
        print("5. Select your Brother QL printer")
        print("6. Install WinUSB driver")
        print(f"7. Visit http://{FLASK_HOST}:{FLASK_PORT}/api/printer/detect to find your printer")
    else:
        print("LINUX:")
        print("1. Install pyusb: pip install pyusb")
        print("2. Add user to groups:")
        print("   sudo usermod -a -G lp $USER")
        print("   sudo usermod -a -G dialout $USER")
        print("3. Create udev rule:")
        print('   echo \'SUBSYSTEM=="usb", ATTR{idVendor}=="04f9", MODE="0666"\' | sudo tee /etc/udev/rules.d/99-brother.rules')
        print("   sudo udevadm control --reload-rules")
        print("4. Reconnect printer or reboot")
        print(f"5. Visit http://{FLASK_HOST}:{FLASK_PORT}/api/printer/detect to find your printer")
    
    print("=" * 60)
    
    # Try to detect printer on startup
    print("\nDetecting connected printers...")
    detected = detect_brother_printer()
    if detected:
        print(f"✓ Found {len(detected)} printer(s)")
        if PRINTER_IDENTIFIER not in detected:
            print(f"⚠️  Warning: Configured printer {PRINTER_IDENTIFIER} not detected.")
            print(f"   Detected printers: {', '.join(detected)}")
            print(f"   Update PRINTER_IDENTIFIER in .env file if needed.")
    else:
        print("⚠️  No printers detected. Check USB connection and drivers.")
    
    print("=" * 60)
    
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG)